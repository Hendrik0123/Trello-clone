import os
import importlib
import subprocess
import sys

# Importiere fehlende Pakete bei Bedarf
def install_and_import(package, import_as=None):
    try:
        return importlib.import_module(import_as or package)
    except ImportError:
        print(f"{package} nicht gefunden. Installiere...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        return importlib.import_module(import_as or package)

# --- Nur externe Pakete prüfen ---
pd = install_and_import("pandas")
openpyxl = install_and_import("openpyxl")
requests = install_and_import("requests")
bs4 = install_and_import("beautifulsoup4", "bs4")
dotenv = install_and_import("python-dotenv", "dotenv")

# Optional: tkinter (manchmal fehlt es auf Linux)
try:
    import tkinter as tk
    from tkinter import ttk
except ImportError:
    print("tkinter nicht gefunden. Unter Linux evtl. installieren mit: sudo apt install python3-tk")
    tk = None
    ttk = None

from pathlib import Path
import re
import json
import tkinter as tk
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings
from dotenv import load_dotenv
import requests
from bs4 import BeautifulSoup

load_dotenv()

# Unterdrücke Warnungen von openpyxl, die nicht relevant sind
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

Aufgaben = "aufgaben.txt"

# Arbeitsverzeichnis (hier: aktuelles Verzeichnis)
VERZEICHNIS = r"Z:\Gruppen\1_NEUE_GRUPPEN"

letzte_meldungen = {}

def finde_ordner_nach_namen(verzeichnis):
    """
    Sucht im angegebenen Verzeichnis nach Ordnern, die Namen innerhalb von Klammern enthalten.
    Der Benutzer wählt anschließend, nach welchem Namen gefiltert werden soll.
    Gibt eine Liste von Tupeln (Ordnername, Typ) zurück.
    """
    import os, re, tkinter as tk
    from tkinter import ttk

    ordner = []
    namen_gefunden = set()
    muster = re.compile(r"\(([^)]+)\)", re.IGNORECASE)

    # Alle Ordner durchsuchen und Namen sammeln
    for name in os.listdir(verzeichnis):
        pfad = os.path.join(verzeichnis, name)
        if os.path.isdir(pfad):
            match = muster.search(name)
            if match:
                gefundene_namen = re.split(r"\+|&|und", match.group(1))
                for einzelner_name in gefundene_namen:
                    normalisierter_name = einzelner_name.strip().lower()
                    if normalisierter_name not in {n.lower() for n in namen_gefunden}:
                        namen_gefunden.add(normalisierter_name.capitalize())

    # Benutzer zur Auswahl eines Namens auffordern
    if not namen_gefunden:
        print("Keine Namen in Klammern gefunden.")
        return []

    # 🔸 "Alle" hinzufügen
    namen_liste = sorted(namen_gefunden)
    namen_liste.insert(0, "Alle")

    # Tkinter-Fenster für die Auswahl
    def auswahl_treffen():
        nonlocal gewaehlter_name
        gewaehlter_name = auswahl_var.get()
        auswahl_fenster.quit()

    gewaehlter_name = None
    auswahl_fenster = tk.Tk()
    auswahl_fenster.title("Namen auswählen")

    tk.Label(auswahl_fenster, text="Bitte einen Namen auswählen:", font=("Arial", 12)).pack(pady=10)

    auswahl_var = tk.StringVar(auswahl_fenster)
    auswahl_var.set(namen_liste[0])  # Standardwert ("Alle")

    dropdown = ttk.Combobox(auswahl_fenster, textvariable=auswahl_var, values=namen_liste, state="readonly")
    dropdown.pack(pady=10)

    tk.Button(auswahl_fenster, text="Auswählen", command=auswahl_treffen).pack(pady=10)

    auswahl_fenster.mainloop()
    auswahl_fenster.destroy()

    if not gewaehlter_name:
        print("Keine Auswahl getroffen. Abbruch.")
        return []

    # Ordner nach dem ausgewählten Namen filtern
    for name in os.listdir(verzeichnis):
        pfad = os.path.join(verzeichnis, name)
        if os.path.isdir(pfad):
            match = muster.search(name)
            if match:
                zwischen = match.group(1)
                # 🔸 Wenn "Alle" gewählt wurde → alles aufnehmen
                if gewaehlter_name == "Alle" or gewaehlter_name.lower() in zwischen.lower():
                    if "+" in zwischen:
                        ordner.append((name, "sekundär"))
                    else:
                        ordner.append((name, "primär"))

    return ordner

def frage_mit_tkinter(Ordnername, frage_text):
    """
    Öffnet ein Tkinter-Fenster mit Buttons für "Ja" und "Nein".
    Gibt die Antwort ("ja" oder "nein") zurück.
    """
    def antwort_ja():
        nonlocal antwort
        antwort = "ja"
        fenster.quit()

    def antwort_nein():
        nonlocal antwort
        antwort = "nein"
        fenster.quit()

    antwort = None
    fenster = tk.Tk()
    fenster.title("Abfrage")

    tk.Label(fenster, text=f"{Ordnername}\n{frage_text}", font=("Arial", 12)).pack(pady=10)

    button_frame = tk.Frame(fenster)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="Ja", command=antwort_ja, width=10).pack(side="left", padx=5)
    tk.Button(button_frame, text="Nein", command=antwort_nein, width=10).pack(side="right", padx=5)

    fenster.mainloop()
    fenster.destroy()

    return antwort

def backup(Ordnername, i):
    regex = r"\(\s*[A-Za-z]{2,}(?:\s*(?:\+|und|&)\s*[A-Za-z]{2,})+\s*\)"
    treffer = re.search(regex, Ordnername)
    eintrag = ws["B6"].value 
    # Prüfung: Wert vorhanden & besteht aus mindestens 2 Wörtern
    if isinstance(eintrag, str) and len(eintrag.strip().split()) >= 2:
        eintrag = True
    else:
        eintrag = False  
    if treffer and eintrag:
        i[1] = datetime.now().date()
    elif treffer and not eintrag:
        return "Backup nicht in Excel eingetragen!"
    elif eintrag and not treffer:
        return "Backup nicht im Ordnernamen eingetragen!"

def termin_GGG(Ordnername, i):
        GGG_Termin = ws["B5"].value.date()
        # Prüfung, ob ein Wert vorhanden ist
        if pd.notna(GGG_Termin):
            i[1] = datetime.now().date()
        else:
            return "Termin für GGG vereinbaren!"
        
def text_warten(Ordnername, i):
    aktuellerpfad = Path(VERZEICHNIS) / Ordnername
    # Prüfen, ob mindestens ein Unterordner existiert
    unterordner = any(p.is_dir() for p in aktuellerpfad.iterdir())
    if unterordner:
        i[1] = datetime.now().date()
    elif not unterordner: 
        heute = datetime.now().date()
        zwei_wochen = timedelta(weeks=2)
        GGG_Termin = ws["B5"].value.date()
        if heute - GGG_Termin > zwei_wochen:
            return "Es sind mehr als 2 Wochen seit dem GGG-Termin vergangen, bitte Initiator:in bzgl. Beschreibung kontaktieren."
        else:
            return "warten auf Text für Homepage / Insta / Presse / Sonstige Kanäle"

def zettel(Ordnername, i):
    frage_text = "Wurde der grüne Zettel Brigitte gegeben?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Grünen Zettel nicht Brigitte gegeben!"

def get_titles_from_url(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
    except Exception as e:
        print(f"Fehler beim Abrufen der URL: {e}")
        return []

    soup = BeautifulSoup(response.text, 'html.parser')
    title_list = []

    # Finde alle divs mit der Klasse 'header'
    headers = soup.find_all("div", class_="header")
    for header in headers:
        # Suche darin nach <h3><a title="...">
        h3 = header.find("h3")
        if h3:
            a_tag = h3.find("a", title=True)
            if a_tag:
                a_tag["title"] = a_tag["title"].replace(",", "")
                a_tag["title"] = a_tag["title"].replace("?", "")
                print(f"Homepage Name: {a_tag['title']}")
                title_list.append(a_tag['title'])

    return title_list

def homepage(Ordnername, i):
    homepage = 0
    excel = 0
    Name = Ordnername.split(" (")[0]
    Gruppen_in_Gruendung = get_titles_from_url(os.getenv("url"))
    if Name in Gruppen_in_Gruendung:
        homepage = 1
    # ist etwas in Zelle B29 eingetragen?
    if pd.notna(df.iloc[27, 1]):
        excel = 1
    if homepage and excel or df.iloc[27, 1] == "-":
        i[1] = datetime.now().date()
    elif not homepage and excel:
        return "Gruppe nicht auf Homepage gefunden! Stimmen Ordnername und Gruppenname überein?"
    elif homepage and not excel:
        return "Gruppe auf Homepage gefunden aber kein Eintrag in Excel vorhanden!"
    else:
        return "Gruppe weder auf Homepage noch in Excel gefunden! Bitte prüfen!"
    

def instagram(Ordnername, i):
    # ist etwas in Zelle B30 eingetragen?
    if pd.notna(df.iloc[29, 1]):
        i[1] = datetime.now().date()
    else:
        return "kein Eintrag in Excel für Instagram vorhanden! (Datum oder '-')"

def presse(Ordnername, i):
    # ist etwas in Zelle B31 eingetragen?
    if pd.notna(df.iloc[30, 1]):
        i[1] = datetime.now().date()
    else:
        return "kein Eintrag in Excel für Pressemitteilung vorhanden! (Datum oder '-')"
            
def interessenten(Ordnername, i): 
    excel_namen = [
    "H4", "H18", "H32", "H46",
    "R4", "R18", "R32", "R46",
    "AB4", "AB18", "AB32", "AB46",
    "AL4", "AL18", "AL32", "AL46",
    "AV4", "AV18", "AV32", "AV46",
    "BF4", "BF18", "BF32", "BF46"
    ]
    anzahl = 0
    heute = datetime.now().date()
    GGG_Termin = ws["B5"].value.date()
    zwei_monate = timedelta(weeks=2)
    for interessent in excel_namen:
        if pd.notna(ws[interessent].value):
            anzahl += 1
    if anzahl >= 4:
        i[1] = datetime.now().date()
    elif anzahl < 4 and heute - GGG_Termin > zwei_monate:
        return f"Das Gründungsgespräch ist mehr als 2 Monate her und es sind weniger als 4 ({anzahl} Personen) Interessent: innen auf der Liste. Initiator: in bzgl. weiterem Vorgehen kontaktieren."
    else:
        return f"Es gibt {anzahl}/4 Interessent: innen für ein erstes Treffen. Warten bis sich min. 4 gemeldet haben."
     

def erstesTreffen(Ordnername, i):   
    # Ist ein Datum in Zelle B21 eingetragen?
    if pd.notna(df.iloc[19, 1]):
        i[1] = datetime.now().date()
    else:
        return "Termin für erstes Treffen vereinbaren!"

def konferenzraum1(Ordnername, i):
    frage_text = f"Wurde der Konferenzraum für das erste Treffen am {df.iloc[19, 1].strftime('%d.%m.%Y')} reserviert?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Konferenzraum nicht reserviert!"

def infoTreffen1(Ordnername, i):
    excel_namen = [
    "H4", "H18", "H32", "H46",
    "R4", "R18", "R32", "R46",
    "AB4", "AB18", "AB32", "AB46",
    "AL4", "AL18", "AL32", "AL46",
    "AV4", "AV18", "AV32", "AV46",
    "BF4", "BF18", "BF32", "BF46",
    ]
    infoCheck = [
    "K4", "K18", "K32", "K46",
    "U4", "U18", "U32", "U46",
    "AE4", "AE18", "AE32", "AE46",
    "AO4", "AO18", "AO32", "AO46",
    "AY4", "AY18", "AY32", "AY46",
    "BI4", "BI18", "BI32", "BI46"
    ]
    nicht_informiert = []
    infoCount = 0
    for interessent in excel_namen:
        if ws[interessent].value is not None:
            if ws[infoCheck[infoCount]].value == False:
                nicht_informiert.append(ws[interessent].value)
        infoCount += 1
    if nicht_informiert == []:
        i[1] = datetime.now().date()
    else:
        return f"wurden folgende Interessent: innen über den ersten Termin am {df.iloc[19, 1].strftime('%d.%m.%Y')} informiert?: \n-{'\n-'.join(nicht_informiert)} \nes ist kein Haken in der Interessiertenliste gesetzt!"
        
def anwesenheit1(Ordnername, i):
    # liegt das erste Treffen in der Vergangenheit?
    if pd.isna(df.iloc[19, 1]):
        return "Termin für erstes Treffen nicht eingetragen!"
    erstesTreffen = df.iloc[19, 1].date()
    heute = datetime.now().date()
    checks = False
    anzahl = False
    anwesenheiten = [
    "M4", "M18", "M32", "M46",
    "W4", "W18", "W32", "W46",
    "AG4", "AG18", "AG32", "AG46",
    "AQ4", "AQ18", "AQ32", "AQ46",
    "BA4", "BA18", "BA32", "BA46",
    "BK4", "BK18", "BK32", "BK46"
    ]
    if erstesTreffen < heute:
        # ist ein Wert in Zelle D21 eingetragen?
        if pd.notna(df.iloc[19, 3]):
            anzahl = True
        # ist mindestens eine Anwesenheit bei den Interessent:innen eingetragen?
        for interessent in anwesenheiten:
            if ws[interessent].value is not None:
                checks = True
                break
        if anzahl and checks:
            i[1] = datetime.now().date()
        elif anzahl and not checks:
            return "Anzahl in Zelle D21 eingetragen aber keine Anwesenheiten abgehakt!"
        elif not anzahl and checks:
            return "Anwesenheiten abgehakt aber keine Anzahl in Zelle D21 eingetragen!"
        else:
            return "weder Anzahl in Zelle D21 eingetragen noch Anwesenheiten abgehakt!"
    else:
        return f"warten bis erstes Treffen am {df.iloc[19, 1].strftime('%d.%m.%Y')} stattgefunden hat!"

def zweitesTreffen(Ordnername, i):    
    # Ist ein Datum in Zelle B21 eingetragen?
    if pd.notna(df.iloc[20, 1]):
        i[1] = datetime.now().date()
    else:
        return "Termin für zweites Treffen vereinbaren!"

        
def konferenzraum2(Ordnername, i):
    frage_text = f"Wurde der Konferenzraum für das erste Treffen am {df.iloc[20, 1].strftime('%d.%m.%Y')} reserviert?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Konferenzraum nicht reserviert!"
        
def infoTreffen2(Ordnername, i):
    excel_namen = [
    "H4", "H18", "H32", "H46",
    "R4", "R18", "R32", "R46",
    "AB4", "AB18", "AB32", "AB46",
    "AL4", "AL18", "AL32", "AL46",
    "AV4", "AV18", "AV32", "AV46",
    "BF4", "BF18", "BF32", "BF46",
    ]
    infoCheck = [
    "K6", "K20", "K34", "K48",
    "U6", "U20", "U34", "U48",
    "AE6", "AE20", "AE34", "AE48",
    "AO6", "AO20", "AO34", "AO48",
    "AY6", "AY20", "AY34", "AY48",
    "BI6", "BI20", "BI34", "BI48"
    ]
    nicht_informiert = []
    infoCount = 0
    for interessent in excel_namen:
        if ws[interessent].value is not None:
            if ws[infoCheck[infoCount]].value == False:
                nicht_informiert.append(ws[interessent].value)
        infoCount += 1
    if nicht_informiert == []:
        i[1] = datetime.now().date()
    else:
        return f"wurden folgende Interessent: innen über den zweiten Termin am {df.iloc[20, 1].strftime('%d.%m.%Y')}  informiert?: \n-{'\n-'.join(nicht_informiert)} \nes ist kein Haken in der Interessiertenliste gesetzt!"
        
def anwesenheit2(Ordnername, i):
    # liegt das zweite Treffen in der Vergangenheit?
    if pd.isna(df.iloc[20, 1]):
        return f"Termin für zweites Treffen nicht eingetragen!"
    zweitesTreffen = df.iloc[20, 1].date()
    heute = datetime.now().date()
    checks = False
    anzahl = False
    anwesenheiten = [
    "M6", "M20", "M34", "M48",
    "W6", "W20", "W34", "W48",
    "AG6", "AG20", "AG34", "AG48",
    "AQ6", "AQ20", "AQ34", "AQ48",
    "BA6", "BA20", "BA34", "BA48",
    "BK6", "BK20", "BK34", "BK48"
    ]
    if zweitesTreffen < heute:
        # ist ein Wert in Zelle D22 eingetragen?
        if pd.notna(df.iloc[20, 3]):
            anzahl = True
        # ist mindestens eine Anwesenheit bei den Interessent:innen eingetragen?
        for interessent in anwesenheiten:
            if ws[interessent].value is not None:
                checks = True
                break
        if anzahl and checks:
            i[1] = datetime.now().date()
        elif anzahl and not checks:
            return "Anzahl in Zelle D22 eingetragen aber keine Anwesenheiten abgehakt!"
        elif not anzahl and checks:
            return "Anwesenheiten abgehakt aber keine Anzahl in Zelle D22 eingetragen!"
        else:
            return "weder Anzahl in Zelle D22 eingetragen noch Anwesenheiten abgehakt!"
    else:
        return f"warten bis zweites Treffen am {df.iloc[20, 1].strftime('%d.%m.%Y')} stattgefunden hat!"
      
def raumsuche(Ordnername, i):
    frage_text = "Hat die Gruppe einen eigenen Raum für weitere Treffen nach dem dritten Termin?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Es muss ein Raum für weitere Treffen gefunden werden!"          
    
def drittesTreffen(Ordnername, i):
        # Ist ein Datum in Zelle B23 eingetragen?
    if pd.notna(df.iloc[21, 1]):
        i[1] = datetime.now().date()
    else:
        return "Termin für drittes Treffen vereinbaren!"

def konferenzraum3(Ordnername, i):
    frage_text = f"Wurde der Konferenzraum für das erste Treffen am {df.iloc[21, 1].strftime('%d.%m.%Y')} reserviert?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Konferenzraum nicht reserviert!"

def infoTreffen3(Ordnername, i):
    excel_namen = [
    "H4", "H18", "H32", "H46",
    "R4", "R18", "R32", "R46",
    "AB4", "AB18", "AB32", "AB46",
    "AL4", "AL18", "AL32", "AL46",
    "AV4", "AV18", "AV32", "AV46",
    "BF4", "BF18", "BF32", "BF46",
    ]
    infoCheck = [
    "K8", "K22", "K36", "K50",
    "U8", "U22", "U36", "U50",
    "AE8", "AE22", "AE36", "AE50",
    "AO8", "AO22", "AO36", "AO50",
    "AY8", "AY22", "AY36", "AY50",
    "BI8", "BI22", "BI36", "BI50"
    ]
    nicht_informiert = []
    infoCount = 0
    for interessent in excel_namen:
        if ws[interessent].value is not None:
            if ws[infoCheck[infoCount]].value == False:
                nicht_informiert.append(ws[interessent].value)
        infoCount += 1
    if nicht_informiert == []:
        i[1] = datetime.now().date()
    else:
        return f"wurden folgende Interessent: innen über den dritten Termin am {df.iloc[21, 1].strftime('%d.%m.%Y')} informiert?: \n-{'\n-'.join(nicht_informiert)} \nes ist kein Haken in der Interessiertenliste gesetzt!"

def fragebogen1(Ordnername, i):
    #Ist ein Wert in Zelle B56 eingetragen?
    if pd.notna(df.iloc[54, 1]):
        i[1] = datetime.now().date()
    else:
        return "Fragebogen an Initiator: in aushändigen!"

def anwesenheit3(Ordnername, i):
    # liegt das dritte Treffen in der Vergangenheit?
    if pd.isna(df.iloc[21, 1]):
        return "Termin für drittes Treffen nicht eingetragen!"
    drittesTreffen = df.iloc[21, 1].date()
    heute = datetime.now().date()
    checks = False
    anzahl = False
    anwesenheiten = [
    "M8", "M22", "M36", "M50",
    "W8", "W22", "W36", "W50",
    "AG8", "AG22", "AG36", "AG50",
    "AQ8", "AQ22", "AQ36", "AQ50",
    "BA8", "BA22", "BA36", "BA50",
    "BK8", "BK22", "BK36", "BK50"
    ]
    if drittesTreffen < heute:
        # ist ein Wert in Zelle D23 eingetragen?
        if pd.notna(df.iloc[21, 3]):
            anzahl = True
        # ist mindestens eine Anwesenheit bei den Interessent:innen eingetragen?
        for interessent in anwesenheiten:
            if ws[interessent].value is not None:
                checks = True
                break
        if anzahl and checks:
            i[1] = datetime.now().date()
        elif anzahl and not checks:
            return "Anzahl in Zelle D23 eingetragen aber keine Anwesenheiten abgehakt!"
        elif not anzahl and checks:
            return "Anwesenheiten abgehakt aber keine Anzahl in Zelle D23 eingetragen!"
        else:
            return "weder Anzahl in Zelle D23 eingetragen noch Anwesenheiten abgehakt!"
    else:
        return f"warten bis drittes Treffen am {df.iloc[21, 1].strftime('%d.%m.%Y')} stattgefunden hat!"

def gruppenbesuch(Ordnername, i):
    frage_text = "Wurde die Gruppe bereits besucht?"
    antwort = frage_mit_tkinter(Ordnername, frage_text)

    if antwort == "ja":
        i[1] = datetime.now().date()
    elif antwort == "nein":
        return "Gruppenbesuch vereinbaren!"

def fragebogen2(Ordnername, i):
    # Ist ein Wert in Zelle B57 eingetragen?
    if pd.notna(df.iloc[55, 1]):
        i[1] = datetime.now().date()
    # Sind seit dem Aushändigen des Fragebogens mehr als 2 Wochen vergangen?
    elif datetime.now().date() - df.iloc[54, 1].date() > timedelta(weeks=2):
        return "Es sind mehr als 2 Wochen seit dem Aushändigen des Fragebogens vergangen, bitte Initiator: in kontaktieren."
    else:
        return "Fragebogen von Initiator: in noch nicht zurückerhalten!"

root = tk.Tk()
root.title("Aufgabenstatus")

frame = tk.Frame(root)
frame.pack(expand=True, fill='both')

baum_pro_gruppe = {}

def update_gui(show_progress=False):
    # Fensterinhalt leeren
    for widget in frame.winfo_children():
        widget.destroy()

    # Wenn Progress-Mode, zeige Progress-Widgets und führe Aktualisierung durch
    if show_progress:
        progress_frame = tk.Frame(frame, padx=10, pady=10)
        progress_frame.pack(expand=True, fill='both')

        progress_label = tk.Label(progress_frame, text="Starte Aktualisierung...", font=("Arial", 12))
        progress_label.pack(pady=(0,10))

        progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=400, mode="determinate")
        progress_bar.pack(pady=(0,10))

        total = len(Gruppen)
        done = 0

        # Prozess: wie in hauptschleife, aber hier mit Fortschritt und ohne zusätzlichen finalen update-Aufruf
        for Gruppe in Gruppen:
            gruppenname = Gruppe[0]
            todo_status_datei = os.path.join(VERZEICHNIS, gruppenname, "todo_status.json")

            # Erstelle todo_status.json falls nötig
            if not os.path.exists(todo_status_datei):
                try:
                    with open(Aufgaben, 'r', encoding='utf-8') as f:
                        zeilen = [zeile.strip() for zeile in f if zeile.strip()]
                    daten = [[zeile, None] for zeile in zeilen]
                    with open(todo_status_datei, 'w', encoding='utf-8') as f:
                        json.dump(daten, f, indent=2, ensure_ascii=False)
                except Exception:
                    daten = []

            # Excel-Datei laden (setzen globaler df/wb/ws für todo_functions)
            datei = list(Path(os.path.join(VERZEICHNIS, gruppenname)).glob('GG Verlauf*.xlsx'))
            if datei:
                pfad = datei[0]
                try:
                    global df, wb, ws
                    df = pd.read_excel(pfad)
                    wb = load_workbook(pfad, data_only=True)
                    ws = wb.active
                except Exception:
                    pass

            # JSON lesen und eine Runde durch die Todos laufen
            try:
                with open(todo_status_datei, 'r', encoding='utf-8') as f:
                    daten = json.load(f)
            except Exception:
                daten = []

            meldung = ""
            for i in daten:
                if i[1] is not None:
                    continue
                else:
                    try:
                        meldung = todo_functions[i[0]](gruppenname, i)
                    except Exception as e:
                        meldung = f"Fehler bei Prüfung: {e}"
                    # JSON nach jeder Funktion speichern
                    try:
                        with open(todo_status_datei, 'w', encoding='utf-8') as f:
                            json.dump(daten, f, indent=2, ensure_ascii=False, default=str)
                    except Exception:
                        pass
                    if meldung:
                        break

            letzte_meldungen[gruppenname] = meldung

            # Fortschritt aktualisieren
            done += 1
            remaining = total - done
            progress_bar['maximum'] = total
            progress_bar['value'] = done
            progress_label.config(text=f"Aktualisiert: {done}/{total} — verbleibend: {remaining}\n {gruppenname}")
            root.update_idletasks()

        # Abschlussanzeige kurz zeigen, dann Progress-Widgets entfernen
        progress_label.config(text="Aktualisierung abgeschlossen.")
        root.update_idletasks()
        root.after(300, lambda: None)
        progress_frame.destroy()

    else:
        # normaler Pfad: hauptschleife wie bisher ausführen, um letzte Meldungen zu erzeugen
        hauptschleife()

    # GUI neu aufbauen (Liste der Gruppen anzeigen)
    for Gruppe in Gruppen:
        gruppenname = Gruppe[0]
        Titel = gruppenname.split(" (")[0]
        MA = f"({gruppenname.split(' (')[1]}" if "(" in gruppenname else ""
        todo_status_datei = os.path.join(VERZEICHNIS, gruppenname, "todo_status.json")
        
        if not os.path.exists(todo_status_datei):
            continue

        # Neuer Frame pro Gruppe
        gruppen_frame = tk.Frame(frame, borderwidth=1, relief="solid", padx=5, pady=5)
        gruppen_frame.pack(side="left", expand=True, fill='both', padx=5, pady=5)

        label = tk.Label(gruppen_frame, text=f"{Titel}\n{MA}", font=("Arial", 12, "bold"))
        label.pack()

        tree = ttk.Treeview(gruppen_frame, columns=("Aufgabe", "Status"), show="headings", height=20)
        tree.heading("Aufgabe", text="Aufgabe")
        tree.heading("Status", text="Status")
        tree.column("Aufgabe", width=150)
        tree.column("Status", width=80)
        tree.pack(expand=True, fill='both')

        try:
            with open(todo_status_datei, 'r', encoding='utf-8') as f:
                daten = json.load(f)
        except Exception:
            daten = []

        for eintrag in daten:
            status = "✅Erledigt" if eintrag[1] else "Offen"
            tree.insert("", "end", values=(eintrag[0], status))

        baum_pro_gruppe[gruppenname] = tree

        # Buttons (binde name + dateipfad als lokale Defaults)
        button_frame = tk.Frame(gruppen_frame)
        button_frame.pack(pady=5)

        ueberspringen_name = gruppenname
        rueckgaengig_name = gruppenname
        ueberspringen_path = todo_status_datei
        rueckgaengig_path = todo_status_datei

        ueberspringen_button = tk.Button(
            button_frame,
            text="Aktuelles ToDo überspringen",
            command=lambda n=ueberspringen_name, p=ueberspringen_path: ueberspringen(n, p)
        )
        ueberspringen_button.pack(side="left", padx=5)

        rueckgaengig_button = tk.Button(
            button_frame,
            text="Vorheriges ToDo rückgängig machen",
            command=lambda n=rueckgaengig_name, p=rueckgaengig_path: rueckgaengig_machen(n, p)
        )
        rueckgaengig_button.pack(side="right", padx=5)

        # Rückmeldung anzeigen, falls vorhanden
        meldung = letzte_meldungen.get(gruppenname, "")
        if meldung:
            meldungs_label = tk.Label(
                gruppen_frame,
                text=meldung,
                fg="green",
                wraplength=180,
                justify="left"
            )
            meldungs_label.pack(pady=(5, 0))

def process_all_groups_with_progress(progress_label, progress_bar):
    """
    Durchläuft alle Gruppen, führt die gleiche Logik wie in hauptschleife aus
    und aktualisiert progress_label und progress_bar nach jeder Gruppe.
    """
    total = len(Gruppen)
    done = 0

    for Gruppe in Gruppen:
        gruppenname = Gruppe[0]
        todo_status_datei = os.path.join(VERZEICHNIS, gruppenname, "todo_status.json")

        # Erstelle todo_status.json falls nötig (wie in hauptschleife)
        if not os.path.exists(todo_status_datei):
            try:
                with open(Aufgaben, 'r', encoding='utf-8') as f:
                    zeilen = [zeile.strip() for zeile in f if zeile.strip()]
                daten = [[zeile, None] for zeile in zeilen]
                with open(todo_status_datei, 'w', encoding='utf-8') as f:
                    json.dump(daten, f, indent=2, ensure_ascii=False)
            except Exception:
                daten = []

        # Excel-Datei laden (wie in hauptschleife)
        datei = list(Path(os.path.join(VERZEICHNIS, gruppenname)).glob('GG Verlauf*.xlsx'))
        if datei:
            pfad = datei[0]
            try:
                # setze globale df/wb/ws damit die todo_functions darauf zugreifen können
                global df, wb, ws
                df = pd.read_excel(pfad)
                wb = load_workbook(pfad, data_only=True)
                ws = wb.active
            except Exception:
                pass

        # JSON lesen und eine Runde durch die Todos laufen (wie in hauptschleife)
        try:
            with open(todo_status_datei, 'r', encoding='utf-8') as f:
                daten = json.load(f)
        except Exception:
            daten = []

        meldung = ""
        for i in daten:
            if i[1] is not None:
                continue
            else:
                try:
                    meldung = todo_functions[i[0]](gruppenname, i)
                except Exception as e:
                    meldung = f"Fehler bei Prüfung: {e}"
                # JSON nach jeder Funktion speichern
                try:
                    with open(todo_status_datei, 'w', encoding='utf-8') as f:
                        json.dump(daten, f, indent=2, ensure_ascii=False, default=str)
                except Exception:
                    pass
                if meldung:
                    break

        letzte_meldungen[gruppenname] = meldung

        # Fortschritt aktualisieren
        done += 1
        remaining = total - done
        progress_bar['maximum'] = total
        progress_bar['value'] = done
        progress_label.config(text=f"Aktualisiert: {done}/{total} — verbleibend: {remaining}")
        # UI-Thread aktualisieren
        root.update_idletasks()

def show_progress_and_refresh():
    # Fensterinhalt leeren
    for w in frame.winfo_children():
        w.destroy()

    # Progress-Frame anlegen
    progress_frame = tk.Frame(frame, padx=10, pady=10)
    progress_frame.pack(expand=True, fill='both')

    progress_label = tk.Label(progress_frame, text="Starte Aktualisierung...", font=("Arial", 12))
    progress_label.pack(pady=(0,10))

    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack(pady=(0,10))

    # Prozess ausführen (synchron, aber UI wird mit update_idletasks aktualisiert)
    process_all_groups_with_progress(progress_label, progress_bar)

    # Kurze Erfolgsmeldung
    progress_label.config(text="Aktualisierung abgeschlossen.")
    root.update_idletasks()
    # kleine Verzögerung damit der User das Ende sieht
    root.after(400)

    # Progress-Frame entfernen und GUI neu aufbauen
    for w in frame.winfo_children():
        w.destroy()
    update_gui()

def ueberspringen(ueberspringen_name, todo_status_datei):
    """Markiert das erste offene ToDo als 'Übersprungen' in der JSON-Datei und zeigt Progress."""
    try:
        with open(todo_status_datei, 'r+', encoding='utf-8') as f:
            daten = json.load(f)
            for eintrag in daten:
                if eintrag[1] is None:
                    eintrag[1] = "Übersprungen"
                    break
            f.seek(0)
            json.dump(daten, f, indent=2, ensure_ascii=False, default=str)
            f.truncate()
    except Exception as e:
        print(f"Fehler beim Überspringen für {ueberspringen_name}: {e}")

    # Fortschritt direkt in update_gui anzeigen (kein doppelter Lauf)
    update_gui(show_progress=True)

def rueckgaengig_machen(rueckgaengig_name, todo_status_datei):
    """Setzt das zuletzt abgeschlossene/übersprungene ToDo zurück auf offen (None) und zeigt Progress."""
    try:
        with open(todo_status_datei, 'r+', encoding='utf-8') as f:
            daten = json.load(f)
            for eintrag in reversed(daten):
                if eintrag[1] is not None:
                    eintrag[1] = None
                    break
            f.seek(0)
            json.dump(daten, f, indent=2, ensure_ascii=False, default=str)
            f.truncate()
    except Exception as e:
        print(f"Fehler beim Rückgängig machen für {rueckgaengig_name}: {e}")

    # Fortschritt direkt in update_gui anzeigen (kein doppelter Lauf)
    update_gui(show_progress=True)

todo_functions = {"Backup Mitarbeiter:in finden": backup,
                  "Mögliche Termine für GGG finden und mit Initiator:in vereinbaren": termin_GGG,
                  "Auf Text für Homepage / Social Media / Pressemitteilung warten": text_warten,
                  "Grünen Zettel Brigitte geben": zettel,
                  "Gruppe auf Homepage inserieren": homepage,
                  "Text an Sabine für Instagram senden": instagram,
                  "Pressemitteilung versenden": presse,
                  "Interessent:innen sammeln": interessenten,
                  "Termin für erstes Treffen vereinbaren": erstesTreffen,
                  "Konferenzraum Reservieren1": konferenzraum1,
                  "Interessent:innen informieren1": infoTreffen1,
                  "Anwesenheiten notieren1": anwesenheit1,
                  "Termin für zweites Treffen vereinbaren": zweitesTreffen,
                  "Konferenzraum Reservieren2": konferenzraum2,
                  "Interessen:innen informieren2": infoTreffen2,
                  "Anwesenheiten notieren2": anwesenheit2,
                  "Initiator:in bei Raumsuche unterstützen": raumsuche,
                  "Termin für drittes Treffen vereinbaren": drittesTreffen,
                  "Konferenzraum Reservieren3": konferenzraum3,
                  "Interessent:innen informieren3": infoTreffen3,
                  "Initiator:in Fragebogen zukommen lassen": fragebogen1,
                  "Anwesenheiten notieren3": anwesenheit3,
                  "Gruppe besuchen": gruppenbesuch,
                  "Fragebogen zurückerhalten und in Datenbank einpflegen": fragebogen2
                }

Gruppen = finde_ordner_nach_namen(VERZEICHNIS)

def hauptschleife():
    for Gruppe in Gruppen:
        gruppenname = Gruppe[0]
        print(gruppenname)
        todo_status_datei = os.path.join(VERZEICHNIS, gruppenname, "todo_status.json")

        if not os.path.exists(todo_status_datei):
            print(f"{gruppenname} todo_status.json nicht vorhanden, wird erstellt")
            with open(Aufgaben, 'r', encoding='utf-8') as f:
                zeilen = [zeile.strip() for zeile in f if zeile.strip()]
            daten = [[zeile, None] for zeile in zeilen]
            with open(todo_status_datei, 'w', encoding='utf-8') as f:
                json.dump(daten, f, indent=2, ensure_ascii=False)
        
        datei = list(Path(os.path.join(VERZEICHNIS, gruppenname)).glob('GG Verlauf*.xlsx'))
        if datei:
            pfad = datei[0]
            global df, wb, ws
            df = pd.read_excel(pfad)
            wb = load_workbook(pfad, data_only=True)
            ws = wb.active
        else:
            print(f"{gruppenname} keine Excel-Datei gefunden!")
            continue

        with open(todo_status_datei, 'r', encoding='utf-8') as f:
            daten = json.load(f)

        meldung = ""  # Rückgabewert der Funktion
        
        for i in daten:
            if i[1] is not None:
                continue
            else:
                meldung = todo_functions[i[0]](gruppenname, i)  # Funktion liefert Text zurück
                with open(todo_status_datei, 'w', encoding='utf-8') as f:
                    json.dump(daten, f, indent=2, ensure_ascii=False, default=str)
                if meldung:
                    break                        

        letzte_meldungen[gruppenname] = meldung  # Merken für die GUI
        
if __name__ == "__main__":
    # Starte initiales Update mit Fortschrittsanzeige nachdem die GUI gestartet ist
    root.after(0, lambda: update_gui(show_progress=True))
    root.mainloop()
